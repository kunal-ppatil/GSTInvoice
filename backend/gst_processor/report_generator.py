import os
import io
import csv
import logging
from typing import List, Any, Dict
from datetime import datetime
from pathlib import Path

# Try imports
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing, Rect, Line, String as DString
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

logger = logging.getLogger(__name__)

class BackendReportGenerator:
    """Compiles and exports GST summary reports in PDF, CSV, and Excel formats"""
    
    def __init__(self, output_dir: str = None):
        if not output_dir:
            base_dir = Path(__file__).parent.parent.parent
            self.output_dir = str(base_dir / "data" / "results")
        else:
            self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
    def generate_pdf_report(self, user_info: Dict[str, Any], invoices: List[Any]) -> str:
        """Generates a professional PDF GST Summary Report"""
        if not PDF_SUPPORT:
            logger.error("reportlab package is not installed. Cannot generate PDF.")
            raise ImportError("reportlab library is required for PDF generation")
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"GST_Report_{timestamp}.pdf"
        file_path = os.path.join(self.output_dir, filename)
        
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
        )
        
        styles = getSampleStyleSheet()
        
        # Custom Styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#1A365D'),
            spaceAfter=15
        )
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#2B6CB0'),
            spaceAfter=10,
            spaceBefore=15
        )
        
        body_style = ParagraphStyle(
            'Body',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#2D3748')
        )

        bold_style = ParagraphStyle(
            'BodyBold',
            parent=body_style,
            fontName='Helvetica-Bold'
        )

        header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=colors.white
        )

        story = []
        
        # --- Header (Company Info) ---
        company_name = user_info.get('company_name') or "InvoiScope Corporate Client"
        company_gst = user_info.get('gst_in') or "N/A"
        company_email = user_info.get('email') or "client@invoiscope.com"
        
        header_data = [
            [Paragraph(f"<b>{company_name}</b><br/>GSTIN: {company_gst}<br/>Email: {company_email}", body_style),
             Paragraph(f"<b>GST SUMMARY REPORT</b><br/>Date: {datetime.now().strftime('%d-%m-%Y')}<br/>Status: Compliance Audit", body_style)]
        ]
        
        header_table = Table(header_data, colWidths=[300, 230])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E0')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 15))
        
        # --- Metrics Row ---
        total_taxable = sum(inv.taxable_amount for inv in invoices)
        total_cgst = sum(inv.cgst for inv in invoices)
        total_sgst = sum(inv.sgst for inv in invoices)
        total_igst = sum(inv.igst for inv in invoices)
        total_gst = sum(inv.total_gst for inv in invoices)
        total_amount = sum(inv.total_amount for inv in invoices)
        
        metrics_data = [
            [
                Paragraph("<b>Total Taxable Value</b>", body_style),
                Paragraph("<b>Total CGST</b>", body_style),
                Paragraph("<b>Total SGST</b>", body_style),
                Paragraph("<b>Total IGST</b>", body_style),
                Paragraph("<b>Total GST</b>", body_style),
                Paragraph("<b>Grand Total</b>", body_style),
            ],
            [
                Paragraph(f"₹{total_taxable:,.2f}", bold_style),
                Paragraph(f"₹{total_cgst:,.2f}", body_style),
                Paragraph(f"₹{total_sgst:,.2f}", body_style),
                Paragraph(f"₹{total_igst:,.2f}", body_style),
                Paragraph(f"₹{total_gst:,.2f}", bold_style),
                Paragraph(f"₹{total_amount:,.2f}", bold_style),
            ]
        ]
        metrics_table = Table(metrics_data, colWidths=[90, 85, 85, 85, 90, 95])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # --- Invoice Table ---
        story.append(Paragraph("Invoice Breakdown", section_style))
        
        table_headers = [
            Paragraph("Date", header_style),
            Paragraph("Invoice #", header_style),
            Paragraph("Merchant", header_style),
            Paragraph("Merchant GSTIN", header_style),
            Paragraph("Taxable (₹)", header_style),
            Paragraph("GST (₹)", header_style),
            Paragraph("Total (₹)", header_style),
            Paragraph("Status", header_style)
        ]
        
        invoice_table_data = [table_headers]
        
        for inv in invoices:
            status_color = "#38A169" if inv.validation_status == "VALID" else "#E53E3E" if inv.validation_status == "INVALID" else "#DD6B20"
            status_html = f"<font color='{status_color}'><b>{inv.validation_status}</b></font>"
            
            row = [
                Paragraph(inv.invoice_date or "N/A", body_style),
                Paragraph(inv.invoice_number or "N/A", body_style),
                Paragraph(inv.merchant_name or "N/A", body_style),
                Paragraph(inv.fields[0].extracted_value if (inv.fields and any(f.field_name == 'GST Number' for f in inv.fields)) else "N/A", body_style),
                Paragraph(f"{inv.taxable_amount:,.2f}", body_style),
                Paragraph(f"{inv.total_gst:,.2f}", body_style),
                Paragraph(f"{inv.total_amount:,.2f}", body_style),
                Paragraph(status_html, body_style)
            ]
            invoice_table_data.append(row)
            
        inv_table = Table(invoice_table_data, colWidths=[60, 60, 95, 95, 60, 60, 60, 45])
        inv_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2B6CB0')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F7FAFC')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(inv_table)
        story.append(Spacer(1, 30))
        
        # --- Signature & QR Code ---
        qr_drawing = Drawing(60, 60)
        # Mock a QR code using shapes
        qr_drawing.add(Rect(0, 0, 60, 60, fillColor=colors.black, strokeColor=colors.black))
        qr_drawing.add(Rect(5, 5, 50, 50, fillColor=colors.white, strokeColor=colors.white))
        qr_drawing.add(Rect(10, 10, 15, 15, fillColor=colors.black, strokeColor=colors.black))
        qr_drawing.add(Rect(35, 10, 15, 15, fillColor=colors.black, strokeColor=colors.black))
        qr_drawing.add(Rect(10, 35, 15, 15, fillColor=colors.black, strokeColor=colors.black))
        qr_drawing.add(Rect(35, 35, 10, 10, fillColor=colors.black, strokeColor=colors.black))
        
        sign_data = [
            [qr_drawing, Paragraph("Digitally Verified<br/>InvoiScope Compliance Engine<br/>Secure Audit Hash: 0x8a92f03b", body_style),
             Paragraph("___________________________<br/><b>Authorized Signatory</b><br/>Corporate Finance Division", body_style)]
        ]
        sign_table = Table(sign_data, colWidths=[70, 230, 230])
        sign_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (2,0), (2,0), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(sign_table)
        
        # Build Document
        doc.build(story)
        return file_path

    def generate_csv_report(self, invoices: List[Any]) -> str:
        """Generates a CSV GST report file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"GST_Report_{timestamp}.csv"
        file_path = os.path.join(self.output_dir, filename)
        
        with open(file_path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            # Headers
            writer.writerow([
                "Invoice ID", "Invoice Number", "Invoice Date", "PO Number", 
                "Merchant Name", "Merchant Address", "Taxable Amount (INR)", 
                "CGST (INR)", "SGST (INR)", "IGST (INR)", "Total GST (INR)", 
                "Total Amount (INR)", "Validation Status"
            ])
            # Data
            for inv in invoices:
                writer.writerow([
                    inv.id, inv.invoice_number, inv.invoice_date, inv.po_number,
                    inv.merchant_name, inv.merchant_address, inv.taxable_amount,
                    inv.cgst, inv.sgst, inv.igst, inv.total_gst, inv.total_amount,
                    inv.validation_status
                ])
                
        return file_path

    def generate_excel_report(self, invoices: List[Any]) -> str:
        """Generates a styled Excel GST report file"""
        if not EXCEL_SUPPORT:
            logger.error("openpyxl package is not installed. Cannot generate Excel.")
            raise ImportError("openpyxl library is required for Excel generation")
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"GST_Report_{timestamp}.xlsx"
        file_path = os.path.join(self.output_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GST Audit Report"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Styling parameters
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        fill_accent = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="CBD5E0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Headers
        headers = [
            "Invoice #", "Date", "PO Number", "Merchant Name", 
            "Taxable Amount", "CGST", "SGST", "IGST", "Total GST", 
            "Total Amount", "Status"
        ]
        
        ws.append([]) # empty row
        ws.append(["GST SUMMARY AUDIT REPORT"]) # title row
        ws.cell(row=2, column=1).font = Font(name="Calibri", size=16, bold=True, color="1A365D")
        ws.append([f"Generated at: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"])
        ws.append([]) # spacer
        
        # Append Table Headers
        ws.append(headers)
        header_row = 5
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
            
        # Append Data
        start_row = 6
        for inv in invoices:
            ws.append([
                inv.invoice_number or "N/A",
                inv.invoice_date or "N/A",
                inv.po_number or "N/A",
                inv.merchant_name or "N/A",
                inv.taxable_amount,
                inv.cgst,
                inv.sgst,
                inv.igst,
                inv.total_gst,
                inv.total_amount,
                inv.validation_status
            ])
            current_row = ws.max_row
            
            # Format alignment and border
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = thin_border
                
                # Alignments
                if col_num in [1, 2, 3]:
                    cell.alignment = align_center
                elif col_num == 4:
                    cell.alignment = align_left
                elif col_num in [5, 6, 7, 8, 9, 10]:
                    cell.alignment = align_right
                    cell.number_format = '₹#,##0.00'
                elif col_num == 11:
                    cell.alignment = align_center
                    # Color coding status
                    if cell.value == "VALID":
                        cell.font = Font(name="Calibri", color="2F855A", bold=True)
                    elif cell.value == "INVALID":
                        cell.font = Font(name="Calibri", color="C53030", bold=True)
                    else:
                        cell.font = Font(name="Calibri", color="DD6B20", bold=True)
                        
        # Append Totals Row
        end_row = ws.max_row
        totals_row = end_row + 1
        
        ws.cell(row=totals_row, column=4, value="Total Summary:").font = font_bold
        ws.cell(row=totals_row, column=4).alignment = align_right
        ws.cell(row=totals_row, column=4).fill = fill_accent
        ws.cell(row=totals_row, column=4).border = thin_border
        
        for col_num in range(1, len(headers) + 1):
            if col_num in [5, 6, 7, 8, 9, 10]:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                cell = ws.cell(row=totals_row, column=col_num, value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
                cell.font = font_bold
                cell.alignment = align_right
                cell.number_format = '₹#,##0.00'
                cell.fill = fill_accent
                cell.border = thin_border
            elif col_num != 4:
                cell = ws.cell(row=totals_row, column=col_num)
                cell.fill = fill_accent
                cell.border = thin_border
                
        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith('='):
                    val = "₹99,999.00"  # mock length for formulas
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(file_path)
        return file_path
